import openpyxl
from openpyxl.styles import Font, PatternFill
import argparse

def write_data_into_excel(txt_filename, xlsx_filename, sort_option):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header = ['Name', 'Surname', 'Age', 'Profession']

    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')
    sheet.append(header)

    header_row = sheet[1]
    for cell in header_row:
        cell.font = bold_font
        cell.fill = yellow_fill

    with open(txt_filename, 'r') as txt_file:
        lines = txt_file.readlines()
        data = []
        for line in lines:
            part = line.split()
            name = part[0]
            surname = part[1]
            age = int(part[2])
            profession = part[3]
            data.append([name, surname, age, profession])

        if sort_option == 'name':
            data.sort(key=lambda x: x[0])
        elif sort_option == 'surname':
            data.sort(key=lambda x: x[1])
        elif sort_option == 'age':
            data.sort(key=lambda x: x[2])
        elif sort_option == 'profession':
            data.sort(key=lambda x: x[3])

        for row_values in data:
            sheet.append(row_values)

            if row_values[2] > 25:
                for cell in sheet[sheet.max_row]:
                    cell.fill = green_fill

    workbook.save(xlsx_filename)
    workbook.close();

def main():
    parser = argparse.ArgumentParser('Process input and output file names, and sorting option.')
    parser.add_argument('-f', '--file', required=True, help='Input file name')
    parser.add_argument('-o', '--output', required=True, help='Output xlsx file name')
    parser.add_argument('-s', '--sort', choices=['name', 'surname', 'age', 'profession'], default='n', help='Sort option (name, surname, age, profession)')

    args = parser.parse_args()
    print(f"Input file name: {args.file}")
    print(f"Output file name: {args.output}")
    print(f"Sort by: {args.sort}")

    write_data_into_excel(args.file, args.output, args.sort)

if  __name__ == "__main__":
    main()
